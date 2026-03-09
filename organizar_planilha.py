from openpyxl.styles import NamedStyle #biblioteca de estilo mui loka pra formata la no excel
import os #biblioteca os, classica
from datetime import datetime, time #datetime que foi o que mais deu problema nessa bagaça
from copy import copy #copy pra copia o estilo da celula modelo e aplicar nas novas celulas, pra manter a formatação
from openpyxl import Workbook, load_workbook #biblioteca pra trabalhar com arquivos excel, o Workbook é pra criar um novo arquivo e o load_workbook é pra carregar um arquivo existente

#classe principal
class ControleDiario:
    placas_validas = ["BPY1A6", "GAA3B19", "GAA3B22", "GAA1C32", "FCJ9C23", "ELE2F27", "FMX9375", "GAA3E02", 
                      "DJL2G36", "FSI5950", "DJM9888", "DJL2F15", "FRH6376", "CDV2573", "BPY9500", "GFR3D53", 
                      "RVL3J48", "RVL7C56", "RVQ1F51", "EWV3A65", "FWM7876", "EFS6E51", "SWH8A30", "DMN4563", "GHZ8170"]
    
    motoristas_validos = ["Adriano", "Alexandre", "Alessandro", "Almir", "Anderson", "André", "Antonio", "Ariel", 
                          "Maria", "Carlos", "Claudio", "Cleber", "Clovis"]
    
    destinos_validos = ["Gruta", "Rodeio", "Bairro Alto", "Funil", "Recanto", "Prainha", "Especiais Geral", "Fazenda", "Creche", "Cidade", "Boa Vista", "Jacarei" ]
    
    meses = ["Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho", "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]

    
#metodo constructor
    def __init__(self, mes="", placa="", data="", motorista="", destino="", saida="", kminicial="", chegada="", kmfinal="", arquivo=None, dados=None):
        self._mes = mes
        self._placa = placa
        self._data = data
        self._motorista = motorista
        self._destino = destino
        self._saida = saida
        self._kminicial = kminicial
        self._chegada = chegada
        self._kmfinal = kmfinal
        self._arquivo = arquivo
        self._dados = dados or []


#50 metodos de get e set
    def mes(self):
        return self._mes

    def mes(self, value):
        self._mes = value

    def placa(self):
        return self._placa

    def placa(self, value):
        self._placa = value


    def data(self):
        return self._data
 
    def data(self, value):
        self._data = value

    def motorista(self):
        return self._motorista
 
    def motorista(self, value):
        self._motorista = value
 
    def destino(self):
        return self._destino
    

    def destino(self, value):
        self._destino = value
  
    def saida(self):
        return self._saida

    def saida(self, value):
        self._saida = value

    def kminicial(self):
        return self._kminicial

    def kminicial(self, value):
        self._kminicial = value
    
    def chegada(self):
        return self._chegada

    def chegada(self, value):
        self._chegada = value

    def kmfinal(self):
        return self._kmfinal

    def kmfinal(self, value):
        self._kmfinal = value

    def arquivo(self):
        return self._arquivo
    

    def arquivo(self, value):
        self._arquivo = value

    def dados(self):
        return self._dados

    def dados(self, value):
        self._dados = value


    def adicionar_motorista(self, nome_motorista): #metodo pra dar umappend na lista de motoristas
        if nome_motorista not in self.motoristas_validos:
            self.motoristas_validos.append(nome_motorista)
            print("✓ Motorista ", nome_motorista, " adicionado à lista.")
        else:
            print("Motorista ", nome_motorista, " já existe na lista.")


#50 coisa diferente pra verificar se os dados inseridos tão certo


    def validar_placa(self):
        if self._placa not in self.placas_validas:
            print("Placas ", self._placa ,"não está na lista de veículos.")
            return False
        return True

    def validar_mes(self):
        if self._mes not in self.meses:
            print("Mês ", self._mes, "inválido. Use: ", ", ".join(self.meses))
            return False
        return True

    def validar_motorista(self):
        if self._motorista not in self.motoristas_validos:
            print("Motorista ",self._motorista, " não encontrado na lista.")
            return False
        return True

    def validar_destino(self):
        if self._destino not in self.destinos_validos:
            print("Destino ", self._destino, " não encontrado na lista.")
            return False
        return True

    def validar_dados(self): #verifica se os dados inseridos sãovalidos
        validacoes = [
            self.validar_mes(),
            self.validar_placa(),
            self.validar_motorista(),
            self.validar_destino()
        ]
        return all(validacoes)

    def criar_aba(self, mes): #metodo pra criar a aba nova caso a entrada não possua uma aba existente
        if mes not in self._arquivo.sheetnames:
            nova_aba = self._arquivo.create_sheet(mes)
            print("✓ Aba ", mes, "' criada com sucesso.")
            return nova_aba
        return self._arquivo[mes]

    def inserir_dados(self): #metodo prtainserir os dados
        print("\n=== Inserir Dados ===")
        
        print("Digite o mês:")
        self._mes = input("").strip()
        
        print("Digite a placa do veículo:")
        self._placa = input("").strip().upper()
        
        print("Digite a data (DD/MM/YYYY):")
        self._data = input("").strip()
        
        print("Digite o nome do motorista:")
        self._motorista = input("").strip()
        
        print("Digite o destino:")
        self._destino = input("").strip()
        
        print("Digite a hora de saída (HH:MM):")
        try:
            saida_input = input("")
            self._saida = saida_input
        except ValueError:
            print("Formato inválido. Use HH:MM")
            self._saida = saida_input
        
        print("Digite o km inicial:")
        self._kminicial = input("").strip()
        
        print("Digite a hora de chegada (HH:MM):")
        try: #try ate q é util, nunca quis usar mas ta ai
            chegada_input = input("")
            self._chegada = chegada_input
        except ValueError:
            print("Formato inválido. Use HH:MM")
            self._chegada = chegada_input
        
        print("Digite o km final:")
        self._kmfinal = input("").strip()

        self._dados = [self._mes, self._placa, self._data, self._motorista, self._destino, 
                       self._saida, self._kminicial, self._chegada, self._kmfinal]
    
    def organizar_planilha(self, nome_arquivo="Planilha de controle Diário.xlsx",): #esse aqui é oprincipal ele faz tudo, LITERALMENTE TUDO, ou seja, ele inseri os dados na planilha
        print("\n=== Processando Planilha ===")

        if not self.validar_dados():
            print("Validação falhou. Operação cancelada.")
            return False

        if not os.path.exists(nome_arquivo):
            print("Arquivo ",nome_arquivo," não encontrado. Criando novo...")
            wb_novo = Workbook()
            wb_novo.save(nome_arquivo)

        self._arquivo = load_workbook(nome_arquivo)
        
        aba_ativa = self._arquivo.active
        celula_modelo = aba_ativa.cell(row=2, column=1)

        aba = self.criar_aba(self._mes)
        linha_inserir = aba.max_row + 1

        for i in range(2, aba.max_row + 1): #isso aqui foi minha genialidade
            if aba.cell(row=i - 1, column=2).value == self._placa and aba.cell(row=i, column=1).value is None and aba.cell(row=i - 1, column=5).value == self._destino:
                linha_inserir = i
                break

        hora_saida = datetime.strptime(self._saida, "%H:%M").time() #isso aqui é coisa do capeta
        hora_chegada = datetime.strptime(self._chegada, "%H:%M").time() #isso tambem, n quero nunca mais mexer nisso

        if "hora_format" not in self._arquivo.named_styles: #capeta ao quadrado
            hora_style = NamedStyle(name="hora_format", number_format="hh:mm")
            self._arquivo.add_named_style(hora_style)

        for col, valor in enumerate(self._dados, start=1): #for pra inserir os dados na planilha
            celula_atual = aba.cell(row=linha_inserir, column=col)
            celula_atual.value = valor

            if celula_modelo.has_style:
                celula_atual.font = copy(celula_modelo.font)
                celula_atual.border = copy(celula_modelo.border)
                celula_atual.fill = copy(celula_modelo.fill)
                celula_atual.alignment = copy(celula_modelo.alignment)


        for col_hora in [6, 8]: #for so pra formata as horas dentro do excel, pq essa desgraça n aceita dado do tipo date da extensão datetime
            celula_hora = aba.cell(row=linha_inserir, column=col_hora)
            celula_hora.value = hora_saida if col_hora == 6 else hora_chegada
            celula_hora.style = "hora_format" 
            celula_hora.alignment = copy(celula_modelo.alignment)

        self._arquivo.save(nome_arquivo) #esse salva o arquivo
        print(f"✓ Dados inseridos com sucesso na linha {linha_inserir} da aba '{self._mes}'")
        return True

if __name__ == "__main__": #main do meu lindo programa

    while True: #prof alexandre iria me matar se visse isso
        print("╔══════════════════════════════╗")
        print("║  Sistema de Controle Diário  ║")
        print("╚══════════════════════════════╝")

        print("\nBem-vindo ao sistema de controle diário de frota! Este programa permite que você registre as atividades diárias dos veículos, incluindo informações como mês, placa, data, motorista, destino, horários de saída e chegada, e quilometragem inicial e final. Os dados serão organizados em uma planilha Excel para fácil consulta e análise futura. Escolha uma opção:")
        print("1. Inserir dados e organizar planilha")
        print("2. Adicionar motorista à lista de motoristas válidos")
        opcao = int(input("Digite o número da opção desejada: "))
        
        controle = ControleDiario()
        if opcao == 1:
            controle.inserir_dados()
            controle.organizar_planilha()
            print("\nDeseja realizar outra operação? (s/n)")
            if input("").strip().lower() == 's':
                continue
            else:
                print("Encerrando o programa. Até a próxima!")
                
                break
        elif opcao == 2:
            nome_motorista = input("Digite o nome do motorista a ser adicionado: ").strip()
            controle.adicionar_motorista(nome_motorista)
            print("\nDeseja realizar outra operação? (s/n)")
            if input("").strip().lower() == 's':
                continue
            else:
                print("Encerrando o programa. Até a próxima!")
                break
